import tkinter as t
from tkinter import ttk, filedialog, messagebox
from os import startfile
import openpyxl as xl
from openpyxl.styles import Font
from parameters import *


class Trust_FY(object):
    def __init__(self, program):
        self.p = program
        self.deletable_widgets = []
        self.run()

    def run(self):
        self.top = t.Toplevel(self.p.root)
        self.top.geometry("+{}+{}".format(SCREEN_POS_RIGHT + 50, SCREEN_POS_DOWN + 100))
        self.top.transient(self.p.root)

        self.main_frame = t.Frame(self.top, bg=SEAGREEN, height=SALE_UHEIGHT * 25, width=1300)
        self.main_frame.pack()
        self.top.update()

        self.menu_frame = t.Frame(self.main_frame, bg=SEAGREEN, height=SALE_UHEIGHT * 25, width=200)
        self.menu_frame.place(x=0, y=0, width=200, height=self.main_frame.winfo_height())
        self.menu_frame.menu_width = 150
        self.top.update()

        self.report_frame = t.Frame(self.main_frame, bg=SEAGREEN, height=SALE_UHEIGHT * 25, width=200)
        self.report_frame.place(x=self.menu_frame.winfo_width(), y=0,
                                             width=self.main_frame.winfo_width() - self.menu_frame.winfo_width(),
                                             height=self.main_frame.winfo_height())
        self.top.update()
        self._canvas_format()
        self._fy_menu1()

    def _canvas_format(self):
        self.canvas = t.Canvas(self.report_frame, bg=SEAGREEN, highlightthickness=0)
        self.canvas.pack(side='left', expand=1, fill='both')

        self.canvas.bind('<Enter>', self._canvas_bound)
        self.canvas.bind('<Leave>', self._canvas_unbound)
        self.top.update()

        self.scrollbar = ttk.Scrollbar(self.report_frame, orient='vertical', command=self.canvas.yview)
        self.scrollbar.pack(side='right', fill='y')

        self.canvas.config(yscrollcommand=self.scrollbar.set)

        self.labels_frame = t.Frame(self.canvas, bg=SEAGREEN, width=self.canvas.winfo_width(), height=self.canvas.winfo_height())
        self.labels_frame.label_width = 125
        self.labels_frame.compiled_labels = []
        self.canvas.create_window((0, 0), window=self.labels_frame, anchor="nw")
        self.top.update()

    def _canvas_bound(self, event):
        self.canvas.bind_all('<MouseWheel>', self._canvas_on_mousewheel)

    def _canvas_unbound(self, event):
        self.canvas.unbind_all('<MouseWheel>')

    def _canvas_on_mousewheel(self, event):
        if self.canvas.bbox("all")[3] > self.report_frame.winfo_height():
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            self.top.update()

    def _fy_menu1(self):
        head = t.Label(self.menu_frame, bg=SEAGREEN, font=font10Cb, text='TRUST VOLUME REPORT')
        head.place(x=0, y=0, width=165, height=SALE_UHEIGHT)

        fy_set1 = {i[2].fy for i in self.p.sales_sorted}
        menu_head = t.StringVar()
        menu_head.set("Start Start FY")
        menu = t.OptionMenu(self.menu_frame, menu_head, (*fy_set1), command=self._fy_menu2)
        menu.place(x=0, y=SALE_UHEIGHT, width=self.menu_frame.menu_width, height=SALE_UHEIGHT * 2)

    def _fy_menu2(self, select):
        self._destroy_labels()
        fy2_set2 = {i[2].fy for i in self.p.sales_sorted if i[2].fy >= select}

        menu_head = t.StringVar()
        menu_head.set("Select End FY")
        menu = t.OptionMenu(self.menu_frame, menu_head, (*fy2_set2),
                            command=lambda selection, sel=select: self._trust_menu(selection, sel=sel))
        menu.place(x=0, y=SALE_UHEIGHT * 4, width=self.menu_frame.menu_width, height=SALE_UHEIGHT * 2)

    def _trust_menu(self, selection, **kwargs):
        self._destroy_labels()
        fy_range = range(int(kwargs['sel']), int(selection) + 1)

        menu_head = t.StringVar()
        menu_head.set("Select Trust")
        menu = t.OptionMenu(self.menu_frame, menu_head, (*TRUST_CODES),
                            command=lambda selection, fy=fy_range: self._compile(selection, fy))
        menu.place(x=0, y=SALE_UHEIGHT * 7, width=self.menu_frame.menu_width, height=SALE_UHEIGHT * 2)

    def _compile(self, selection, fy_range):
        self._destroy_labels()
        heads = ['SALE', 'FISCAL YEAR', 'TRUST', 'ACRES', 'MBF', 'DNR REVENUE', 'TRUST REVENUE', 'TOTAL REVENUE']
        x = 0
        fy_head = t.Label(self.labels_frame, bg=SEAGREEN, font=font10Cb, anchor='w',
                          text=f'FISCAL YEARS:  {min(fy_range)} - {max(fy_range)}')
        fy_head.place(x=x, y=0, width=200, height=SALE_UHEIGHT)
        self.deletable_widgets.append(fy_head)
        for i in heads:
            label = t.Label(self.labels_frame, bg=SEAGREEN, font=font10Cb, anchor='w', text=i)
            label.place(x=x, y=SALE_UHEIGHT, width=self.labels_frame.label_width, height=SALE_UHEIGHT)
            self.deletable_widgets.append(label)
            x += self.labels_frame.label_width

        self._labels(selection, fy_range, heads)

    def _export_excel_button(self, headers, display_list, totals):
        export_list = [headers] + display_list + [totals]
        button = t.Button(self.menu_frame, font=font10Cb, text='Export to Excel')
        button['command'] = lambda el=export_list: self._export_to_excel(el)
        button.place(x=0, y=SALE_UHEIGHT * 10, width=self.menu_frame.menu_width, height=SALE_UHEIGHT)
        self.deletable_widgets.append(button)

    def _export_to_excel(self, export_list):
        save = filedialog.asksaveasfilename(initialdir=self.p.path, title="Save Excel File",
                                              filetypes=(("Excel Files", "*.xlsx*"), ("All Files", "*.*")), parent=self.top)

        if save != '':
            save = self._file_extension_check(save, '.xlsx')
            wb = xl.Workbook()
            ws = wb.active
            ws.title = f'Trust {export_list[1][2]}'
            b_font = Font(bold=True)
            for i, row in enumerate(export_list):
                for j, col in enumerate(row):
                    if i == 0:
                        cell = ws.cell(row=i+1, column=j+1)
                        cell.font = b_font
                        cell.value = col
                    else:
                        cell = ws.cell(row=i + 1, column=j + 1)
                        cell.value = col
                        if j >= 5:
                            #EXCEL ACCOUNTING FORMAT
                            cell.number_format = r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'
            wb.save(save)
            q = messagebox.askyesno('Completed', f'Completed Exporting to Excel\n{save}\n\nWould you like to open the file?',
                                    parent=self.top)
            if q:
                startfile(save)

    def _labels(self, selection, fy_range, headers):
        display_list, totals = self._get_sales(selection, fy_range)

        y = SALE_UHEIGHT * 2
        if len(display_list) == 0:
            label = t.Label(self.labels_frame, bg=SEAGREEN, font=font10Cb, anchor='w', text='No Trusts within FY Range')
            label.place(x=0, y=y, width=200, height=SALE_UHEIGHT)
            self.deletable_widgets.append(label)
        else:
            self._export_excel_button(headers, display_list, totals)
            for sale in display_list:
                x = 0
                for i, dat in enumerate(sale):
                    if i == 3 or i == 4:
                        text = self._format_number_with_commas(dat)
                    else:
                        text = dat
                    label = t.Label(self.labels_frame, bg=SEAGREEN, font=font10Cb, anchor='w', text=text)
                    label.place(x=x, y=y, width=self.labels_frame.label_width, height=SALE_UHEIGHT)
                    self.deletable_widgets.append(label)
                    x += self.labels_frame.label_width
                y += SALE_UHEIGHT
            y += SALE_UHEIGHT

            x = 0
            for i, tot in enumerate(totals):
                if i == 3 or i == 4:
                    text = self._format_number_with_commas(tot)
                else:
                    text = tot
                label = t.Label(self.labels_frame, bg=SEAGREEN, font=font10Cb, anchor='w', text=text)
                label.place(x=x, y=y, width=self.labels_frame.label_width, height=SALE_UHEIGHT)
                self.deletable_widgets.append(label)
                x += self.labels_frame.label_width
        self.top.update()
        self.labels_frame['height'] = y + (SALE_UHEIGHT * 3)
        self.canvas['scrollregion'] = self.canvas.bbox("all")
        self.canvas.update()

    def _get_sales(self, selection, fy_range):
        temp = []
        totals = ['TOTALS', '', selection, 0, 0, 0, 0, 0]
        for sale in self.p.sales_sorted:
            if sale[2].fy in fy_range:
                if selection in sale[2].trusts:
                    temp.append([sale[1], sale[2].fy, selection,
                                 int(round(sale[2].trusts[selection][ACRES], 0)),
                                 int(round(sale[2].trusts[selection][MBF], 0)),
                                 self._format_currency(sale[2].trusts[selection][MBF] * sale[2].est_val_mbf * DNR_REVENUE[selection]),
                                 self._format_currency(sale[2].trusts[selection][MBF] * sale[2].est_val_mbf * TRUST_REVENUE[selection]),
                                 self._format_currency(sale[2].trusts[selection][MBF] * sale[2].est_val_mbf)])

                    totals[3] += int(round(sale[2].trusts[selection][ACRES], 0))
                    totals[4] += int(round(sale[2].trusts[selection][MBF], 0))
                    totals[5] += sale[2].trusts[selection][MBF] * sale[2].est_val_mbf * DNR_REVENUE[selection]
                    totals[6] += sale[2].trusts[selection][MBF] * sale[2].est_val_mbf * TRUST_REVENUE[selection]
                    totals[7] += sale[2].trusts[selection][MBF] * sale[2].est_val_mbf

        for i in range(5, 8):
            totals[i] = self._format_currency(totals[i])

        return temp, totals

    def _format_number_with_commas(self, value):
        text = str(int(round(value, 0)))
        if len(text) > 3:
            text_list = [i for i in text]
            text_list.insert(-3, ',')
            return ''.join([i for i in text_list])
        else:
            return text

    def _format_currency(self, value):
        val_list = [i for i in str(round(value, 2))]
        if '.' not in val_list:
            add_to = ['.', '0', '0']
            for i in add_to:
                val_list.append(i)
        else:
            if len(val_list[-(len(val_list) - val_list.index('.')):]) < 3:
                val_list.append('0')
        temp = [i for i in reversed(val_list)]
        added = 0
        for i in range(3, len(val_list)):
            if i != 3 and i % 3 == 0:
                temp.insert(i + added, ',')
                added += 1
        return '${}'.format(''.join([i for i in reversed(temp)]))

    def _file_extension_check(self, filename, extension):
        check = ''.join([filename[i] for i in range(-len(extension), 0)])
        if check != extension:
            filename += extension
        return filename

    def _destroy_labels(self):
        if len(self.deletable_widgets) > 0:
            for i in range(len(self.deletable_widgets)):
                self.deletable_widgets.pop(0).destroy()



